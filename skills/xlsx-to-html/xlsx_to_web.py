#!/usr/bin/env python3
"""
Excel to HTML Converter

Converts xlsx/xls files to standalone HTML that preserves the visual structure.
Includes search/filter functionality for easy browsing.

Usage:
    python xlsx_to_web.py input.xlsx
    python xlsx_to_web.py input.xlsx --output-dir ./docs
    python xlsx_to_web.py ./excel_folder/
"""

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Union

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


class ExcelToHtml:
    """Converts Excel files to HTML preserving visual structure."""

    def __init__(self, input_path: Union[str, Path], output_dir: Union[str, Path] = None):
        self.input_path = Path(input_path)
        self.output_dir = Path(output_dir) if output_dir else self.input_path.parent
        self.workbook = None
        self.data: Dict[str, List[Dict[str, Any]]] = {}

    def load(self) -> "ExcelToHtml":
        """Load the Excel file."""
        if not self.input_path.exists():
            raise FileNotFoundError(f"File not found: {self.input_path}")

        self.workbook = openpyxl.load_workbook(self.input_path, data_only=True)
        self._extract_data()
        return self

    def _serialize_value(self, value: Any) -> Any:
        """Convert cell values to displayable types."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return value

    def _extract_data(self) -> None:
        """Extract data from all sheets."""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                self.data[sheet_name] = []
                continue

            # First row as headers
            headers = [str(h) if h is not None else f"Col_{i}" for i, h in enumerate(rows[0], 1)]

            # Remaining rows as data
            sheet_data = []
            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                row_dict = {h: self._serialize_value(v) for h, v in zip(headers, row)}
                sheet_data.append(row_dict)

            self.data[sheet_name] = sheet_data

    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        if not isinstance(text, str):
            text = str(text) if text else ""
        return (text.replace("&", "&amp;").replace("<", "&lt;")
                    .replace(">", "&gt;").replace('"', "&quot;"))

    def to_html(self) -> str:
        """Convert to standalone HTML with search functionality."""
        total_rows = sum(len(rows) for rows in self.data.values())

        html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self._escape_html(self.input_path.stem)}</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Arial, sans-serif;
            margin: 0; padding: 20px; background: #f5f5f5; color: #333;
        }}
        h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-bottom: 5px; }}
        .meta {{ color: #666; font-size: 13px; margin-bottom: 20px; }}
        .search-box {{ margin-bottom: 20px; }}
        .search-box input {{
            width: 100%; max-width: 400px; padding: 10px 15px; font-size: 16px;
            border: 2px solid #ddd; border-radius: 25px; outline: none;
        }}
        .search-box input:focus {{ border-color: #3498db; }}
        .tabs {{ display: flex; flex-wrap: wrap; gap: 5px; margin-bottom: 0; }}
        .tab {{
            padding: 10px 20px; background: #fff; border: none;
            border-radius: 5px 5px 0 0; cursor: pointer; font-size: 14px;
        }}
        .tab:hover {{ background: #e0e0e0; }}
        .tab.active {{ background: #3498db; color: white; }}
        .sheet {{
            display: none; background: #fff; border-radius: 0 8px 8px 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1); overflow: hidden;
        }}
        .sheet.active {{ display: block; }}
        .table-wrap {{ overflow-x: auto; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
        th {{
            background: #3498db; color: white; padding: 12px 15px;
            text-align: left; position: sticky; top: 0; cursor: pointer;
            white-space: nowrap;
        }}
        th:hover {{ background: #2980b9; }}
        td {{ padding: 10px 15px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f8f9fa; }}
        tr.hidden {{ display: none; }}
        .highlight {{ background: #fff3cd; }}
        .row-count {{ padding: 10px 15px; background: #f8f9fa; color: #666; font-size: 13px; }}
        .empty {{ padding: 40px; text-align: center; color: #999; }}
    </style>
</head>
<body>
    <h1>{self._escape_html(self.input_path.stem)}</h1>
    <p class="meta">Source: {self._escape_html(self.input_path.name)} | {total_rows} rows | Converted: {datetime.now().strftime("%Y-%m-%d %H:%M")}</p>
    <div class="search-box">
        <input type="text" id="search" placeholder="Search..." oninput="filter()">
    </div>
    <div class="tabs">'''

        # Tabs
        for i, name in enumerate(self.data.keys()):
            active = " active" if i == 0 else ""
            html += f'<button class="tab{active}" onclick="showSheet(\'{name}\')">{self._escape_html(name)}</button>'

        html += '</div>'

        # Sheets
        for i, (name, rows) in enumerate(self.data.items()):
            active = " active" if i == 0 else ""
            sheet_id = name.replace(" ", "_").replace("'", "")

            if not rows:
                html += f'<div id="sheet_{sheet_id}" class="sheet{active}"><div class="empty">No data</div></div>'
                continue

            headers = list(rows[0].keys())
            html += f'<div id="sheet_{sheet_id}" class="sheet{active}"><div class="table-wrap"><table id="tbl_{sheet_id}">'
            html += '<thead><tr>'
            for j, h in enumerate(headers):
                html += f'<th onclick="sort(\'tbl_{sheet_id}\',{j})">{self._escape_html(h)}</th>'
            html += '</tr></thead><tbody>'

            for row in rows:
                html += '<tr>'
                for h in headers:
                    val = row.get(h, "")
                    html += f'<td>{self._escape_html(str(val))}</td>'
                html += '</tr>'

            html += f'</tbody></table></div><div class="row-count">Showing {len(rows)} of {len(rows)} rows</div></div>'

        html += '''
    <script>
        function showSheet(name) {
            document.querySelectorAll('.sheet').forEach(s => s.classList.remove('active'));
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.getElementById('sheet_' + name.replace(/ /g,'_').replace(/'/g,'')).classList.add('active');
            event.target.classList.add('active');
        }
        function filter() {
            const q = document.getElementById('search').value.toLowerCase();
            document.querySelectorAll('table').forEach(tbl => {
                let count = 0;
                tbl.querySelectorAll('tbody tr').forEach(row => {
                    const match = !q || row.textContent.toLowerCase().includes(q);
                    row.classList.toggle('hidden', !match);
                    if (match) count++;
                    row.querySelectorAll('td').forEach(td => {
                        td.classList.toggle('highlight', q && td.textContent.toLowerCase().includes(q));
                    });
                });
                const rc = tbl.parentElement.parentElement.querySelector('.row-count');
                if (rc) rc.textContent = `Showing ${count} of ${tbl.querySelectorAll('tbody tr').length} rows`;
            });
        }
        function sort(id, col) {
            const tbl = document.getElementById(id);
            const tbody = tbl.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const asc = tbl.dataset.sortCol === String(col) && tbl.dataset.sortDir === 'asc';
            rows.sort((a, b) => {
                const av = a.cells[col]?.textContent || '';
                const bv = b.cells[col]?.textContent || '';
                const an = parseFloat(av), bn = parseFloat(bv);
                if (!isNaN(an) && !isNaN(bn)) return asc ? bn - an : an - bn;
                return asc ? bv.localeCompare(av) : av.localeCompare(bv);
            });
            tbl.dataset.sortCol = col;
            tbl.dataset.sortDir = asc ? 'desc' : 'asc';
            rows.forEach(r => tbody.appendChild(r));
        }
    </script>
</body>
</html>'''
        return html

    def save(self) -> Path:
        """Save HTML to output directory."""
        self.output_dir.mkdir(parents=True, exist_ok=True)
        output_path = self.output_dir / f"{self.input_path.stem}.html"
        output_path.write_text(self.to_html(), encoding="utf-8")
        return output_path


def main():
    parser = argparse.ArgumentParser(description="Convert Excel files to HTML")
    parser.add_argument("input", help="Excel file or directory")
    parser.add_argument("--output-dir", "-o", help="Output directory (default: same as input)")
    args = parser.parse_args()

    input_path = Path(args.input)

    if input_path.is_dir():
        files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))
        if not files:
            print(f"No Excel files found in {input_path}")
            return
        for f in files:
            print(f"Converting: {f.name}")
            out = ExcelToHtml(f, args.output_dir).load().save()
            print(f"  -> {out}")
    elif input_path.is_file():
        print(f"Converting: {input_path.name}")
        out = ExcelToHtml(input_path, args.output_dir).load().save()
        print(f"  -> {out}")
    else:
        print(f"Error: {input_path} not found")
        sys.exit(1)


if __name__ == "__main__":
    main()
